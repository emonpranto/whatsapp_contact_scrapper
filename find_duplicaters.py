import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

df1 = pd.read_excel("generative Ai (NLP).xlsx")
df2 = pd.read_excel("Gen AI - Webinarnumbers.xlsx")

column_name = "Phone_Numbers"


duplicates = df1[df1[column_name].isin(df2[column_name])]

wb = load_workbook("New_Generative_Ai(NLP).xlsx")
ws = wb.active


red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


for index, row in duplicates.iterrows():
    cell = ws.cell(row=index + 2, column=df1.columns.get_loc(column_name) + 1) 
    cell.fill = red_fill


output_file = "Duplicates.xlsx"
wb.save(output_file)

print(f"Duplicate numbers have been marked in red. Saved as {output_file}")